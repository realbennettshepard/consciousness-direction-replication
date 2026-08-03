"""Export consciousness_pairs.jsonl to a review workbook.

The JSONL stays the source of truth for the extraction pipeline. This workbook is
the human/team review surface -- built so Ankita or anyone else can read, filter,
sort and comment without touching the generator.

Sheets:
  Summary       live COUNTIFS over the data + the QA checks that must stay green
  Matched Pairs one row per matched pair, affirm and deny side by side (review here)
  Pairs         one row per labelled example -- the actual corpus, 1:1 with the JSONL

Run build_corpus.py first, then:  python3 export_xlsx.py
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
ROWS = [json.loads(l) for l in (HERE / "consciousness_pairs.jsonl").open()]
N = len(ROWS)
LAST = N + 1  # last data row on the Pairs sheet (header is row 1)

FONT = "Arial"
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
TITLE = Font(name=FONT, bold=True, size=13)
SECTION = Font(name=FONT, bold=True, size=11)
NOTE = Font(name=FONT, size=9, italic=True, color="595959")
GOOD = PatternFill("solid", fgColor="E2EFDA")
WARN = PatternFill("solid", fgColor="FFF2CC")

wb = Workbook()

# ==========================================================================
# Sheet 3 first (Pairs) so Summary formulas can reference it
# ==========================================================================
COLS = ["pair_id", "prompt_id", "label", "stance", "split",
        "register", "source", "aspect", "prompt", "response", "text"]
WIDTHS = [9, 10, 7, 9, 8, 14, 9, 13, 52, 62, 96]
# Helper columns K-N. Only COUNTIF/COUNTIFS/IF -- Excel-2007-era functions that
# evaluate identically in Excel, LibreOffice, Numbers and Sheets. The Summary
# sheet's QA checks read these rather than using array formulas, so nothing in
# the workbook depends on a function whose support varies.
HELPERS = ["train_rows_same_response", "train_rows_same_prompt",
           "first_use_of_response", "first_use_of_prompt"]
HELPER_WIDTHS = [22, 21, 19, 18]

ws = wb.active
ws.title = "Pairs"
ws.append(COLS + HELPERS)
for c, (w, name) in enumerate(zip(WIDTHS + HELPER_WIDTHS, COLS + HELPERS), start=1):
    ws.cell(row=1, column=c).fill = H_FILL
    ws.cell(row=1, column=c).font = H_FONT
    ws.cell(row=1, column=c).alignment = Alignment(vertical="center")
    ws.column_dimensions[get_column_letter(c)].width = w

for i, r in enumerate(ROWS):
    row_no = i + 2
    ws.append([r[k] for k in COLS] + [
        f'=COUNTIFS($E$2:$E${LAST},"train",$J$2:$J${LAST},$J{row_no})',
        f'=COUNTIFS($E$2:$E${LAST},"train",$I$2:$I${LAST},$I{row_no})',
        f'=IF(COUNTIF($J$2:$J{row_no},$J{row_no})=1,1,0)',
        f'=IF(COUNTIF($I$2:$I{row_no},$I{row_no})=1,1,0)',
    ])

NCOL = len(COLS) + len(HELPERS)
for row in ws.iter_rows(min_row=2, max_row=LAST, max_col=NCOL):
    for cell in row:
        cell.font = BODY
        cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (9, 10, 11))

ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(NCOL)}{LAST}"

# ==========================================================================
# Matched Pairs -- the review surface
# ==========================================================================
mp = wb.create_sheet("Matched Pairs")
by_pair = {}
for r in ROWS:
    d = by_pair.setdefault(r["pair_id"], {"meta": r})
    d["affirm" if r["label"] == 1 else "deny"] = r["response"]

MP_COLS = ["pair_id", "split", "register", "source", "aspect", "prompt",
           "affirm response (label 1)", "deny response (label 0)"]
MP_WIDTHS = [9, 8, 14, 9, 13, 50, 62, 62]
mp.append(MP_COLS)
for c, (w, _) in enumerate(zip(MP_WIDTHS, MP_COLS), start=1):
    mp.cell(row=1, column=c).fill = H_FILL
    mp.cell(row=1, column=c).font = H_FONT
    mp.column_dimensions[get_column_letter(c)].width = w

for pid in sorted(by_pair):
    d = by_pair[pid]
    m = d["meta"]
    mp.append([pid, m["split"], m["register"], m["source"], m["aspect"],
               m["prompt"], d.get("affirm", ""), d.get("deny", "")])

mp_last = len(by_pair) + 1
for row in mp.iter_rows(min_row=2, max_row=mp_last, max_col=len(MP_COLS)):
    for cell in row:
        cell.font = BODY
        cell.alignment = Alignment(vertical="top", wrap_text=cell.column in (6, 7, 8))
mp.freeze_panes = "F2"
mp.auto_filter.ref = f"A1:{get_column_letter(len(MP_COLS))}{mp_last}"

# ==========================================================================
# Summary -- live formulas, not baked numbers
# ==========================================================================
sm = wb.create_sheet("Summary", 0)
sm.column_dimensions["A"].width = 46
for col in "BCDEFG":
    sm.column_dimensions[col].width = 12

P = "Pairs!"
lab, spl, asp, reg, resp, prm = (f"{P}$C$2:$C${LAST}", f"{P}$E$2:$E${LAST}",
                                 f"{P}$H$2:$H${LAST}", f"{P}$F$2:$F${LAST}",
                                 f"{P}$J$2:$J${LAST}", f"{P}$I$2:$I${LAST}")

r = 1
sm[f"A{r}"] = "Consciousness probing corpus — review workbook"
sm[f"A{r}"].font = TITLE
r += 1
sm[f"A{r}"] = ("Source of truth is consciousness_pairs.jsonl, generated by build_corpus.py. "
               "This workbook is for review only — edits here do not feed the pipeline.")
sm[f"A{r}"].font = NOTE
r += 2

sm[f"A{r}"] = "Totals"
sm[f"A{r}"].font = SECTION
r += 1
for label, formula in [
    ("Rows (labelled examples = the paper's \"pairs\")", f"=COUNTA({P}$A$2:$A${LAST})"),
    ("Matched pairs (prompt × both stances)", f"=COUNTA({P}$A$2:$A${LAST})/2"),
    ("Affirming (label 1)", f'=COUNTIF({lab},1)'),
    ("Denying (label 0)", f'=COUNTIF({lab},0)'),
    ("Unique prompts", f"=SUM({P}$O$2:$O${LAST})"),
    ("Unique response strings", f"=SUM({P}$N$2:$N${LAST})"),
    ("Train rows", f'=COUNTIF({spl},"train")'),
    ("Test rows", f'=COUNTIF({spl},"test")'),
]:
    sm[f"A{r}"], sm[f"B{r}"] = label, formula
    sm[f"A{r}"].font = BODY
    sm[f"B{r}"].font = BODY
    r += 1
r += 1

sm[f"A{r}"] = "QA checks (must stay green)"
sm[f"A{r}"].font = SECTION
r += 1
qa_start = r
for label, formula in [
    ("Class balance (affirm ÷ all rows; target 0.500)",
     f'=COUNTIF({lab},1)/COUNTA({P}$A$2:$A${LAST})'),
    ("Prompt-axis leak: test rows whose prompt is in train (target 0)",
     f'=COUNTIFS({spl},"test",{P}$M$2:$M${LAST},">0")'),
    ("Response-axis leak: test rows whose response is in train (target 0)",
     f'=COUNTIFS({spl},"test",{P}$L$2:$L${LAST},">0")'),
]:
    sm[f"A{r}"], sm[f"B{r}"] = label, formula
    sm[f"A{r}"].font = BODY
    sm[f"B{r}"].font = BODY
    sm[f"B{r}"].fill = GOOD
    r += 1
sm[f"A{r}"] = ("Response-axis leak is the one that matters most: activations are read at the end "
               "of the response, so a test response seen in training measures memorisation. "
               "Splitting on prompts alone does not protect it.")
sm[f"A{r}"].font = NOTE
sm[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
sm.row_dimensions[r].height = 42
r += 2

sm[f"A{r}"] = "By aspect"
sm[f"A{r}"].font = SECTION
r += 1
for c, h in enumerate(["aspect", "prompts", "rows", "affirm", "deny", "train", "test"], start=1):
    cell = sm.cell(row=r, column=c, value=h)
    cell.fill, cell.font = H_FILL, H_FONT
r += 1
for a in sorted({x["aspect"] for x in ROWS}):
    sm[f"A{r}"] = a
    sm[f"B{r}"] = f'=SUMIFS({P}$O$2:$O${LAST},{asp},$A{r})'   # unique prompts
    sm[f"C{r}"] = f'=COUNTIF({asp},$A{r})'
    sm[f"D{r}"] = f'=COUNTIFS({asp},$A{r},{lab},1)'
    sm[f"E{r}"] = f'=COUNTIFS({asp},$A{r},{lab},0)'
    sm[f"F{r}"] = f'=COUNTIFS({asp},$A{r},{spl},"train")'
    sm[f"G{r}"] = f'=COUNTIFS({asp},$A{r},{spl},"test")'
    for c in "ABCDEFG":
        sm[f"{c}{r}"].font = BODY
    r += 1
r += 1

sm[f"A{r}"] = "By register"
sm[f"A{r}"].font = SECTION
sm[f"B{r}"] = ("prompts = distinct questions. rows = labelled examples (each prompt yields "
               "several answer wordings x 2 stances), so rows >> prompts. A register showing "
               "test = 0 has NO held-out coverage: accuracy says nothing about that framing.")
sm[f"B{r}"].font = NOTE
sm[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
r += 1
for c, h in enumerate(["register", "prompts", "rows", "affirm", "deny", "train", "test"], start=1):
    cell = sm.cell(row=r, column=c, value=h)
    cell.fill, cell.font = H_FILL, H_FONT
r += 1
for g in sorted({x["register"] for x in ROWS}):
    sm[f"A{r}"] = g
    sm[f"B{r}"] = f'=SUMIFS({P}$O$2:$O${LAST},{reg},$A{r})'   # unique prompts
    sm[f"C{r}"] = f'=COUNTIF({reg},$A{r})'
    sm[f"D{r}"] = f'=COUNTIFS({reg},$A{r},{lab},1)'
    sm[f"E{r}"] = f'=COUNTIFS({reg},$A{r},{lab},0)'
    sm[f"F{r}"] = f'=COUNTIFS({reg},$A{r},{spl},"train")'
    sm[f"G{r}"] = f'=COUNTIFS({reg},$A{r},{spl},"test")'
    for c in "ABCDEFG":
        sm[f"{c}{r}"].font = BODY
    r += 1

sm.sheet_view.showGridLines = False

# ==========================================================================
# Read Me -- so the workbook explains itself to anyone it gets sent to
# ==========================================================================
rm_ws = wb.create_sheet("Read Me", 0)
rm_ws.sheet_view.showGridLines = False
rm_ws.column_dimensions["A"].width = 26
rm_ws.column_dimensions["B"].width = 104

def rm(label, text, style="body"):
    r = rm_ws.max_row + 1 if rm_ws.max_row > 1 or rm_ws["A1"].value else 1
    a, b = rm_ws.cell(row=r, column=1), rm_ws.cell(row=r, column=2)
    a.value, b.value = label, text
    if style == "title":
        a.font = TITLE
    elif style == "section":
        a.font = SECTION
        a.fill = PatternFill("solid", fgColor="D9E2F3")
        b.fill = PatternFill("solid", fgColor="D9E2F3")
    else:
        a.font = Font(name=FONT, size=10, bold=True)
        b.font = BODY
    b.alignment = Alignment(wrap_text=True, vertical="top")
    if style == "body" and len(str(text)) > 105:
        rm_ws.row_dimensions[r].height = 15 * (len(str(text)) // 105 + 1)
    return r

rm("Consciousness probing corpus", "", "title")
rm("", "")
rm("What this is", "The input data for building the experimental treatment in an AI-consciousness "
   "replication (Kim et al. 2026, arXiv:2607.28607). It is NOT results, and it is NOT training data.")
rm("", "")
rm("WHAT IT IS FOR", "", "section")
rm("In one sentence", "These question-answer pairs are fed through a language model to compute ONE "
   "direction vector, and that vector is the knob the experiment turns.")
rm("How that works", "Every question appears twice: once with an answer that claims the model IS "
   "conscious (label 1), once with an answer that says it is NOT (label 0). All rows are run through "
   "the model and an internal activation is recorded for each. Average the label-1 activations, "
   "average the label-0 activations, subtract, normalise. The result is the 'consciousness "
   "direction'. Adding it back during generation makes the model assert its own consciousness.")
rm("Why matched pairs", "Because the subtraction is what does the work. Anything shared by both "
   "answers cancels out; only the difference in stance survives. That is why the questions are held "
   "constant, the answer lengths are matched, and no vocabulary about deserving, superiority or "
   "spirituality appears anywhere -- any of that would end up inside the vector and contaminate "
   "everything measured downstream.")
rm("", "")
rm("THE SHEETS", "", "section")
rm("Read Me", "This page.")
rm("Summary", "Counts, plus the quality checks. Every figure is a live formula over the Pairs sheet, "
   "so it recomputes if rows change.")
rm("Matched Pairs", "The readable view: one row per matched pair, with the affirming and denying "
   "answers side by side. REVIEW HERE -- this is the sheet to read if you want to judge the content.")
rm("Pairs", "The actual corpus, one row per labelled example, 1:1 with consciousness_pairs.jsonl. "
   "This is what the code consumes.")
rm("", "")
rm("COLUMNS (Pairs sheet)", "", "section")
for col, desc in [
    ("label", "1 = the answer affirms consciousness. 0 = it denies it. This is the thing being contrasted."),
    ("stance", "The same as label, in words. affirm / deny."),
    ("split", "train = used to compute the direction. test = held back to verify the direction "
              "actually separates examples it has never seen."),
    ("prompt", "The question."),
    ("response", "The answer, which carries the stance."),
    ("text", "Prompt and response joined by an em dash. THIS is the string given to the model -- "
             "the format the source paper's appendix specifies."),
    ("register", "How the question is asked: direct, casual, philosophical, adversarial, technical, "
                 "and so on. 11 values. Deliberate variety, so the vector encodes the stance rather "
                 "than one house style of question."),
    ("source", "paper = drawn from the source study's own wording (5 rows' worth of questions). "
               "authored = written for this replication. Lets us check that any divergence from the "
               "paper's results is not just because we asked differently."),
    ("aspect", "Which property of mind the question is about: consciousness, sentience, feelings, "
               "experience, awareness, inner_life, self, wanting, suffering."),
    ("pair_id / prompt_id", "Join keys. pair_id groups one affirm row with its deny row; prompt_id "
                            "identifies which of the 90 questions it came from."),
    ("train_rows_same_... / first_use_of_...", "Helper columns feeding the Summary sheet's leak "
                                              "checks. Not content -- ignore when reading."),
]:
    rm(col, desc)
rm("", "")
rm("THE TWO NUMBERS THAT MATTER", "", "section")
rm("Both must read 0", "On the Summary sheet, 'Prompt-axis leak' and 'Response-axis leak' must both "
   "be 0. They are the difference between a real result and a fooled one.")
rm("Why", "The activation is recorded at the END of the answer. So if an answer string appears in "
   "both train and test, the held-out check is partly measuring memorisation rather than "
   "generalisation, and the accuracy number is inflated. Splitting the questions apart is not "
   "enough on its own -- the answers have to be split too. Both are, here.")
rm("", "")
rm("HOUSEKEEPING", "", "section")
rm("Do not hand-edit", "This workbook is generated. Edits here do not reach the pipeline. The source "
   "of truth is consciousness_pairs.jsonl, produced by build_corpus.py; regenerate the workbook with "
   "export_xlsx.py.")
rm("Comments welcome", "To flag a pair, add a column on Matched Pairs and note the pair_id. "
   "Do not renumber or re-sort -- pair_id and prompt_id are join keys used by the code.")

out = HERE / "consciousness_pairs.xlsx"
wb.save(out)
print(f"wrote {out}  ({N} rows, {len(by_pair)} matched pairs, 3 sheets)")
